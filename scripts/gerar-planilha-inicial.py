# Gera a planilha inicial do bot financeiro (sheets/bot-financeiro.xlsx)
# com as 8 abas e dados iniciais de EXEMPLO (genéricos — substitua pelos seus).
# Depois: fazer upload no Google Drive e abrir como Google Sheets.
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

CRIADO_EM = "2026-01-01"
HEADER_FONT = Font(name="Arial", bold=True, color="FFFFFF")
HEADER_FILL = PatternFill("solid", start_color="2E5E37")
BODY_FONT = Font(name="Arial")

# Dados de EXEMPLO. Tudo aqui é placeholder genérico — troque pelos seus próprios
# contas fixas, dicionário de classificação, metas e configurações.
ABAS: dict[str, tuple[list[str], list[list]]] = {
    "Lançamentos": (
        ["data_competencia", "data_original", "descricao", "titulo", "valor",
         "categoria", "tipo", "origem", "status", "id_meta"],
        [],
    ),
    "Contas Fixas": (
        ["nome", "dia_vencimento", "valor_esperado", "ativo"],
        [
            ["Internet", 8, 120.00, "sim"],
            ["Energia", 8, 250.00, "sim"],
            ["Condomínio", 5, 800.00, "sim"],
            # exemplo de conta semanal (cobrada toda sexta)
            ["Diarista", "sexta-feira", 400.00, "sim"],
            ["Gás", 11, 100.00, "sim"],
            ["Academia", 5, 150.00, "sim"],
            ["Serviço Mensal", 5, 200.00, "sim"],
        ],
    ),
    "Contas Variáveis": (
        ["nome", "categoria", "observacao"],
        [],
    ),
    "Dicionário": (
        ["descricao_original", "categoria_mapeada", "origem", "criado_em"],
        # Conta corrente (chave: campo Título) — exemplos genéricos
        [["FORNECEDOR DIARISTA", "Diarista", "conta", CRIADO_EM],
         ["ADMINISTRADORA CONDOMINIO", "Condomínio", "conta", CRIADO_EM],
         ["FORNECEDOR GAS", "Gás", "conta", CRIADO_EM],
         ["OPERADORA TELECOM", "Internet", "conta", CRIADO_EM],
         ["ORGAO PUBLICO IMPOSTO", "Meta: IPTU", "conta", CRIADO_EM],
         # Transferência entre contas próprias: parser resolve pela direção
         # (entrada -> Pagamento, saída -> Retirada); literal nunca vai a Lançamentos
         ["TITULAR DA CONTA", "Pagamento/Retirada", "conta", CRIADO_EM],
         # Cartão de crédito (chave: campo Descrição) — exemplos genéricos
         ["MERCADO EXEMPLO", "Supermercado", "cartao", CRIADO_EM],
         ["RESTAURANTE EXEMPLO", "Alimentação", "cartao", CRIADO_EM],
         ["APP DELIVERY", "Alimentação", "cartao", CRIADO_EM],
         ["STREAMING EXEMPLO", "Streams", "cartao", CRIADO_EM],
         ["CIA AEREA EXEMPLO", "Meta: Viagem", "cartao", CRIADO_EM],
         ["LOJA ONLINE EXEMPLO", "Compras", "cartao", CRIADO_EM]],
    ),
    "Categorias": (
        ["nome", "tipo", "ativo"],
        [["Internet", "fixa", "sim"],
         ["Energia", "fixa", "sim"],
         ["Condomínio", "fixa", "sim"],
         ["Diarista", "fixa", "sim"],
         ["Gás", "fixa", "sim"],
         ["Academia", "fixa", "sim"],
         ["Serviço Mensal", "fixa", "sim"],
         ["Supermercado", "variável", "sim"],
         ["Alimentação", "variável", "sim"],
         ["Streams", "variável", "sim"],
         ["Compras", "variável", "sim"],
         ["Outros", "variável", "sim"],
         ["Depósito Pessoa B", "entrada", "sim"],
         ["Depósito Pessoa A", "entrada", "sim"],
         ["Bônus", "entrada", "sim"],
         ["Juros", "entrada", "sim"],
         ["Pagamento", "entrada", "sim"],
         ["Retirada", "entrada", "sim"],
         ["Outros (entrada)", "entrada", "sim"]],
    ),
    "Metas": (
        ["nome", "orcamento_total", "valor_acumulado", "prazo", "status", "criado_em"],
        [["Meta Viagem", "", 0, "", "ativa", CRIADO_EM],
         ["Meta Casa", "", 0, "", "ativa", CRIADO_EM],
         ["Meta Eletrodoméstico", "", 0, "", "ativa", CRIADO_EM],
         ["Meta Plantas", "", 0, "", "ativa", CRIADO_EM],
         ["Meta IPTU", "", 0, "", "ativa", CRIADO_EM],
         ["Meta Evento", "", 0, "", "ativa", CRIADO_EM]],
    ),
    "Config": (
        ["chave", "valor"],
        [["cartao_vencimento_dia", "10"],
         ["telegram_chat_id", "PREENCHER_NO_SETUP"],
         ["diarista_nome", "PREENCHER_NO_SETUP"],
         ["diarista_valor", "400.00"],
         ["gemini_model", "gemini-2.0-flash-preview"]],
    ),
    "Log": (
        ["timestamp", "acao", "entidade", "valor_anterior", "valor_novo", "origem"],
        [],
    ),
}


def main() -> None:
    wb = Workbook()
    wb.remove(wb.active)
    for nome, (header, linhas) in ABAS.items():
        ws = wb.create_sheet(nome)
        ws.append(header)
        for col in range(1, len(header) + 1):
            cell = ws.cell(row=1, column=col)
            cell.font = HEADER_FONT
            cell.fill = HEADER_FILL
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col)].width = max(14, len(header[col - 1]) + 4)
        for linha in linhas:
            ws.append(linha)
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.font = BODY_FONT
        ws.freeze_panes = "A2"
    destino = Path(__file__).resolve().parent.parent / "sheets" / "bot-financeiro.xlsx"
    destino.parent.mkdir(exist_ok=True)
    wb.save(destino)
    print(f"OK: {destino}")


if __name__ == "__main__":
    main()
